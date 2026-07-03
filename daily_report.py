#!/usr/bin/env python3
"""情报日报 — 抓取多平台热点，生成 Excel 表格，微信通知"""

import subprocess
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATE = datetime.now().strftime("%Y-%m-%d")
FILENAME = f"情报日报_{DATE}.xlsx"

# ========== 数据抓取 ==========

def bb(site_cmd):
    """运行 bb-browser site 命令，返回 data"""
    cmd = f"bb-browser site {site_cmd}"
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=60)
    if result.returncode != 0:
        return {}
    try:
        return json.loads(result.stdout)
    except:
        return {}


def fetch_all():
    print("抓取中...")
    zhihu = bb("zhihu/hot")
    v2ex = bb("v2ex/hot")
    kr36 = bb("36kr/newsflash")
    weibo = bb("weibo/hot")
    xueqiu = bb("xueqiu/hot-stock 10")
    bili = bb("bilibili/search 'AI agent'")
    print("抓取完成")
    return zhihu, v2ex, kr36, weibo, xueqiu, bili


# ========== Excel 生成 ==========

def style_sheet(ws, title, col_widths):
    """统一样式"""
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_widths))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    for col, (name, width) in enumerate(col_widths, 1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 22

    return 3  # 数据起始行


def fill_rows(ws, start_row, data_rows, col_count):
    """填入数据"""
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

    for i, row in enumerate(data_rows):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=start_row + i, column=j, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = alt_fill
        ws.row_dimensions[start_row + i].height = 36


def build_excel(zhihu, v2ex, kr36, weibo, xueqiu, bili):
    wb = Workbook()
    wb.remove(wb.active)

    # --- 封面 ---
    ws0 = wb.create_sheet("日报封面")
    ws0.merge_cells("A1:D1")
    c = ws0.cell(row=1, column=1, value=f"情报日报 | {DATE}")
    c.font = Font(name="微软雅黑", size=20, bold=True, color="2F5496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 50

    items_data = []
    for item in zhihu.get("data", {}).get("items", [])[:5]:
        items_data.append([item["rank"], item["title"], f'{item["heat"]}', item["answer_count"]])
    if items_data:
        style_sheet(ws0, "", [("排名", 6), ("标题", 60), ("热度", 14), ("回答数", 10)])
        fill_rows(ws0, 3, items_data, 4)

    # --- 知乎热榜 ---
    items = zhihu.get("data", {}).get("items", [])
    if not items:
        items = zhihu.get("items", [])
    ws1 = wb.create_sheet("知乎热榜")
    cols = [("排名", 6), ("标题", 70), ("热度", 14), ("回答数", 10)]
    start = style_sheet(ws1, "知乎热榜", cols)
    rows = [[i["rank"], i["title"], i.get("heat", ""), i.get("answer_count", "")]
            for i in items]
    fill_rows(ws1, start, rows, 4)

    # --- V2EX ---
    items = v2ex.get("data", {}).get("topics", [])
    if not items:
        items = v2ex.get("topics", [])
    ws2 = wb.create_sheet("V2EX")
    cols = [("标题", 60), ("节点", 10), ("回复", 8)]
    start = style_sheet(ws2, "V2EX 热帖", cols)
    rows = [[i["title"], i.get("node", ""), i.get("replies", "")]
            for i in items]
    fill_rows(ws2, start, rows, 3)

    # --- 36氪 ---
    items = kr36.get("items", [])
    ws3 = wb.create_sheet("36氪快讯")
    cols = [("标题", 50), ("内容摘要", 80)]
    start = style_sheet(ws3, "36氪 快讯", cols)
    rows = [[i["title"], (i.get("description") or "")[:200]] for i in items]
    fill_rows(ws3, start, rows, 2)

    # --- 微博热搜 ---
    items = weibo.get("items", [])
    ws4 = wb.create_sheet("微博热搜")
    cols = [("排名", 6), ("关键词", 30), ("热度", 14), ("分类", 12)]
    start = style_sheet(ws4, "微博热搜", cols)
    rows = [[i["rank"], i["word"], i.get("hot_value", ""), i.get("category", "")]
            for i in items]
    fill_rows(ws4, start, rows, 4)

    # --- 雪球热股 ---
    items = xueqiu.get("items", [])
    ws5 = wb.create_sheet("雪球热股")
    cols = [("股票名", 15), ("涨跌幅", 10)]
    start = style_sheet(ws5, "雪球人气股票", cols)
    rows = [[i["name"], i.get("changePercent", "")] for i in items]
    fill_rows(ws5, start, rows, 2)

    # --- B站搜索 ---
    items = bili.get("videos", [])
    ws6 = wb.create_sheet("B站AI视频")
    cols = [("标题", 60), ("UP主", 20), ("播放量", 12)]
    start = style_sheet(ws6, "B站 AI Agent 热门视频", cols)
    rows = [[i["title"], i.get("author", ""), i.get("play", "")]
            for i in items]
    fill_rows(ws6, start, rows, 3)

    wb.save(FILENAME)
    return FILENAME


# ========== 微信通知 ==========

def notify_wechat(filename):
    webhook = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=0283b64b-0f0a-44d5-b72c-7f050acdbd6e"
    import requests
    msg = f"情报日报已生成\n文件: {filename}\n日期: {DATE}"
    requests.post(webhook, json={"msgtype": "text", "text": {"content": msg}})


if __name__ == "__main__":
    data = fetch_all()
    fname = build_excel(*data)
    print(f"已生成: {fname}")
    notify_wechat(fname)
    print("微信已通知")
