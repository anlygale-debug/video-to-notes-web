#!/usr/bin/env python3
"""生成 UI 灵感库两大方向项目深度对比 Excel（含链接）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ===== 样式 =====
hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_font = Font(name="Arial", size=10, bold=True, color="2F5496")
c_font = Font(name="Arial", size=9)
link_font = Font(name="Arial", size=8, color="0563C1", underline="single")
b_font = Font(name="Arial", size=9, bold=True)
g_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

def style_hdr(ws, row, n):
    for c in range(1, n+1):
        cl = ws.cell(row=row, column=c)
        cl.font = hdr_font; cl.fill = hdr_fill; cl.alignment = center; cl.border = thin

def style_row(ws, row, n, is_sub=False):
    for c in range(1, n+1):
        cl = ws.cell(row=row, column=c)
        cl.font = sub_font if is_sub else c_font
        cl.fill = sub_fill if is_sub else PatternFill()
        cl.alignment = wrap; cl.border = thin

# ═══════════════════════════════════════
# Sheet 1: 方向一 — 储存 UI 灵感的工具
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "方向一-储存UI灵感"

h1 = ["项目名称", "链接", "类型", "Stars", "核心能力", "代码/设计片段",
      "标签分类", "导出格式", "AI兼容", "自托管", "费用", "匹配度", "推荐"]
for c, t in enumerate(h1, 1): ws1.cell(row=1, column=c, value=t)
style_hdr(ws1, 1, len(h1))

dir1 = [
    ["═══ 代码片段+书签管理类 ═══",]*13,
    ["ByteBox", "github.com/nickxla/bytebox", "代码片段\n管理器", "⭐200+", "语法高亮+看板视图\n玻璃拟态UI\n100+语言代码着色", "★★★★★\nShiki高亮\n代码卡片", "标签+看板\n6种主题", "JSON/HTML", "中等\n手动复制", "✅Docker", "MIT", "存代码好\n不存截图", "备选"],
    ["Seahorse", "github.com/SSBun/Seahorse", "macOS\n书签App", "新", "双击复制自动保存\nAI自动打标签\nMarkdown预览", "★★★☆☆\n文本片段\n无代码高亮", "AI自动标签\nSF Symbols", "系统书签", "低\nmacOS原生", "✅原生", "MIT", "Mac独占\n快速捕获", "备选"],
    ["Faved", "github.com/denho/faved", "书签\n管理器", "持续", "嵌套标签+重复检测\n自动元数据\n导入Raindrop/Pocket", "★★☆☆☆\n存链接\n非代码", "★★★★★\n嵌套标签\n深度层级", "JSON/HTML\n/CSV", "低", "✅PWA", "MIT", "标签最强\n灵感目录", "备选"],
    ["Bookmrk", "github.com/sarvan-2187/Bookmrk", "看板\n书签", "新", "Trello式看板\n3种主题(极简/粗野)\n键盘优先", "★★☆☆☆\n存链接", "看板列+卡片\n拖拽排序", "本地JSON", "低", "✅纯前端", "MIT", "视觉友好\n看板直观", "备选"],

    ["═══ UI组件专用存储 ═══",]*13,
    ["Uiverse\nGalaxy", "uiverse.io\ngithub.com/uiverse-io/galaxy", "社区UI库", "3000+\n组件", "3000+开源UI元素\nCSS+Tailwind双格式\n社区审核", "★★★★★\n完整源码\n一键复制", "分类+标签\nButtons/Cards\n等类别", "复制代码", "★★★★★\n纯CSS/TW\nAI直接可用", "N/A\n平台托管", "MIT", "组件最多\n非自己存", "★★★☆☆"],
    ["Storybook\nv10", "github.com/storybookjs/storybook", "组件文档\n行业标准", "⭐85K+", "组件隔离开发+文档\n视觉测试+无障碍\nMCP Server(AI可读)", "★★★★★\n代码+故事+MDX\n完整文档", "组件层级树\nStory分类", "静态站点", "★★★★★\nMCP原生\nAI直接查", "✅Pages", "MIT", "个人组件库\n标答,有曲线", "★★★★☆"],
    ["Story UI", "github.com/southleft/story-ui", "AI故事\n生成器", "v4.16", "NL→Storybook故事\n多框架+多LLM\n自我修复", "★★★★★\nAI生成故事\nTS验证", "组件story\n层级", "Storybook\n集成", "★★★★★\n多LLM兼容\nDS目录导入", "✅", "MIT", "Storybook\nAI增强", "★★★★☆"],
    ["ReUI", "github.com/keenthemes/ReUI", "shadcn\n模式库", "⭐2.6K+", "966+生产级模式\nRadix+Base UI双支持\n所有5种shadcn风格", "★★★★★\n完整源码\nshadcn注册表", "组件+模式\n26+类别", "注册表JSON\n+TSX源码", "★★★★★\nshadcn CLI\nAI完美兼容", "✅", "MIT", "模式最全\n学组件组合", "★★★★☆"],
    ["Uikki", "npmjs.com/package/uikki", "CLI\n组件库", "新", "零依赖+自托管\n受shadcn/ui启发\n组件直接写入src/", "★★★★★\n源码复制到项目\n完全掌控", "组件名索引", "源码文件", "★★★★★\nshadcn风格\nAI友好", "✅", "MIT", "理念好\n组件少", "★★★☆☆"],

    ["═══ AI生成型灵感库 ═══",]*13,
    ["UI Syntax", "ui-syntax.com", "AI生成\n组件", "日增3\n组件", "Gemini 24/7自动生成\nPlaywright视觉QA\nTailwind+React", "★★★★★\nAI生成代码+截图\n自动化流水线", "标签分类\n每日更新", "代码复制", "★★★★★\nAI原生\n自动QA", "N/A", "免费", "灵感自动\n不需手动", "★★★☆☆"],
    ["Component\nGallery", "component.gallery", "组件参考", "开源", "93+设计系统对照\n无障碍ARIA指引\n跨系统参考", "★★★★☆\n设计规范对照\n实现参考", "组件+系统名\n对照索引", "文档", "★★★★☆\n结构化参考\nAI可学", "✅", "开源", "设计决策参考\n非代码存储", "★★★☆☆"],

    ["═══ 设计灵感专用 ═══",]*13,
    ["Style\nGallery", "github.com/changeroa/StyleGallery", "CSS模式", "个人\n项目", "最小化CSS布局模式\n每个解决一个空间问题", "★★★★★\n纯HTML+CSS源码\n极简可移植", "按布局类型", "源码文件", "★★★★★\n纯CSS片段\nAI直接应用", "✅", "开源", "小巧精准\n布局灵感", "★★★☆☆"],
]

row = 2
for d in dir1:
    is_sub = str(d[0]).startswith("═══")
    for c, v in enumerate(d[:13], 1):
        cl = ws1.cell(row=row, column=c, value=v)
        if c == 2 and not is_sub:
            cl.font = link_font
    style_row(ws1, row, 13, is_sub=is_sub)
    row += 1

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 28
for col in ['C','D','E','F','G','H','I','J','K','L','M']:
    ws1.column_dimensions[col].width = 20
ws1.freeze_panes = "C2"

# ═══════════════════════════════════════
# Sheet 2: 方向二 — 学习 UI 结构的项目
# ═══════════════════════════════════════
ws2 = wb.create_sheet("方向二-学习UI结构")

h2 = ["项目名称", "链接", "类型", "Stars", "核心内容", "覆盖量",
      "输出格式", "AI可读性", "Obsidian兼容", "活跃度", "费用", "匹配度", "推荐"]
for c, t in enumerate(h2, 1): ws2.cell(row=1, column=c, value=t)
style_hdr(ws2, 1, len(h2))

dir2 = [
    ["═══ DESIGN.md 生态系统 ═══",]*13,
    ["VoltAgent/\nawesome-\ndesign-md", "github.com/VoltAgent/awesome-design-md", "DESIGN.md\n集合", "⭐91K+\nTop150", "58+品牌完整设计系统\n每份9节:色板/字体/组件\n布局/阴影/Do&Don't", "58+品牌\nStripe/Apple\nNotion等", "Markdown\n+preview\n.html", "★★★★★\n专为AI设计\n含Agent提示", "★★★★★\n纯MD\n直接拖入", "🔥爆火\n持续新增", "MIT", "★★★★★\n基因档案\n模板", "🥇首选"],
    ["Asahina\nMafuuyuu/\nawesome-\ndesign-md", "github.com/AsahinaMafuuyuu/awesome-design-md", "DESIGN.md\n社区分支", "活跃", "按行业分类:AI平台/DevTools\n金融科技/汽车/消费品牌\n更细粒度分类", "类似\nVoltAgent\n+不同视角", "Markdown\n(DESIGN\n.md)", "★★★★★\n同格式", "★★★★★\n同格式", "活跃", "MIT", "★★★★★\n分类互补\n参考", "🥇首选"],
    ["designmd\n.sh", "designmd.sh", "CLI\n注册表", "新\nVoltAgent\n维护", "公共注册表\nnpx一键安装DESIGN.md\n支持14种AI工具", "索引GitHub\n所有\nDESIGN.md", "npx自动\n安装到\n项目根", "★★★★★\n14种AI\n原生兼容", "★★★★☆\nMD文件\n可搬入", "2026.6\n活跃", "免费", "★★★★☆\n安装方便\n少搬运", "推荐"],
    ["TypeUI", "typeui.sh", "CLI创作\n+注册表", "新\nMIT", "生成SKILL.md+DESIGN.md\n50+手作设计技能\n交互式生成+随机化", "50+设计\n技能文件", "SKILL.md\n+DESIGN\n.md双格式", "★★★★★\n双格式\nClaude原生", "★★★★★\n可生成\nObsidian笔记", "2026.4\n活跃", "MIT\nPro可选", "★★★★☆\n创作新DS\n偏制作", "推荐"],
    ["design-\nbites", "npmjs.com/package/design-bites", "CLI设计\n片段库", "v0.2.0\n270+网站", "270+网站→DESIGN.md\n零依赖+npx运行\n按域名模糊搜索", "270+网站\nlinear.app\nstripe.com", "DESIGN\n.md写入\n项目根", "★★★★★\n按需下载\n无冗余", "★★★★★\n直接拿MD", "增长中", "MIT", "★★★★★\n量大速获\n灵感", "🥇首选"],
    ["mddesign", "npmjs.com/package/mddesign", "全生命\n周期CLI", "v0.7.0", "完整流程:init→validate\n→publish→render\n5规则校验", "用户发布\n填充", "DESIGN\n.md+预览", "★★★★☆\n发布+验证\n质量保证", "★★★★☆\nMD输出\n+在线预览", "新工具", "免费", "★★★☆☆\n偏发布端", "备选"],

    ["═══ 设计系统学习资源 ═══",]*13,
    ["distill-\ndesign", "github.com/tototomato1457/distill-design", "网站蒸馏\n工具", "新", "任意网站→DESIGN.md\n替代Gemini手动分析\n自动提取设计系统", "任意网站\n自动蒸馏", "DESIGN\n.md", "★★★★★\n自动生成\nAI可读DS", "★★★★★\n直接产出\nObsidian笔记", "新工具", "开源", "★★★★★\n网页→基因\n自动化", "🥇首选"],
    ["brad\ntraversy/\ndesign-\nresources", "github.com/bradtraversy/design-resources-for-developers", "设计资源\n大全", "⭐经典", "27分类:UI图形/字体/颜色\n图标/CSS框架/AI工具\n400+资源", "400+资源\n27分类", "链接索引\nMarkdown", "★★★☆☆\n资源目录\n需加工", "★★★★☆\n链接列表\n可做索引", "经典活跃", "免费", "★★★☆☆\n资源目录\n非本体", "备选"],
    ["dalisoft/\nawesome-ui-\nlibraries", "github.com/dalisoft/awesome-ui-libraries", "UI库大全", "⭐经典", "灵感网站+40+React库\n25+shadcn生态+AI组件\nMCP工具", "数百项目\n海量分类", "链接索引\nMarkdown", "★★★☆☆\n目录导航\n需自己探索", "★★★★☆\n索引文档\n可存", "活跃", "MIT", "★★★☆☆\n工具发现\n入口", "备选"],

    ["═══ UX模式+设计模式学习 ═══",]*13,
    ["ux-patterns\n-for-\ndevelopers", "github.com/thedaviddias/ux-patterns-for-developers\nuxpatterns.dev", "UX模式\n文档", "⭐活跃", "UX模式完整文档\n最佳实践+无障碍+实现\n真实案例+可复用组件", "Navigation\nForms\nModals等", "文档+代码\nMarkdown", "★★★★★\n模式文档+代码\nAI可学习", "★★★★★\n纯MD文档\n完美兼容", "活跃", "开源", "★★★★☆\nUX模式学习\n非视觉灵感", "推荐"],
    ["UXonFly\nMCP", "github.com/Phanikondru/uxonfly-mcp", "UX规则\nMCP", "2026.4\n新发布", "Modern SaaS美学\n9种UX模式+3种流程\n7条设计原则", "导航/模态\n表单/加载\n等9种模式", "MCP规则\nAI直接读", "★★★★★\nMCP服务器\nAI原生调用", "★★★☆☆\nMCP格式\n非纯MD", "活跃", "MIT", "★★★★☆\nAI实时查\n设计规则", "推荐"],
    ["Raven\nMCP", "github.com/rhinocap/raven-mcp", "设计智能\nMCP", "新", "8层设计知识体系\n模式+内容+服务设计\nStripe/Linear token", "注册/定价\n导航/表单\n仪表板等", "MCP服务器\n规则+token", "★★★★★\n8层知识\nAI深度集成", "★★★☆☆\nMCP格式\n需转换", "活跃", "开源", "★★★★☆\n最全面AI\n设计知识库", "推荐"],
    ["github-\nelements", "github.com/github/github-elements", "GitHub\nWeb组件", "⭐经典\nGitHub官方", "17+Web Components源码\nclipboard-copy/relative-time\n等生产级组件剖析", "17+组件\n逐一学习", "Web组件\n源码+文档", "★★★★☆\n源码+规范\n学组件拆分", "★★★★☆\n代码+文档\n可存", "经典维护", "MIT", "★★★☆☆\n学习GitHub\n组件思想", "备选"],

    ["═══ 开源设计系统参考 ═══",]*13,
    ["facebook/\nastryx", "github.com/facebook/astryx", "Meta\n设计系统", "⭐Meta\n90+组件", "AI就绪+可完全定制\n13,000+内部App验证\nStyleX构建", "90+组件\n企业级", "React组件\n+文档", "★★★★☆\nAI原生设计\n可学结构", "★★★☆☆\nReact代码\n非纯MD", "Meta维护", "MIT", "★★★☆☆\n学企业级\nDS架构", "备选"],
    ["Geeklego", "github.com/geekyants/geeklego", "AI原生\n设计系统", "持续增长\n81组件", "3层token架构\n6个AI技能\n完整文档+Storybook", "81组件\n完整系统", "组件+token\n+Storybook", "★★★★★\n6个AI技能\n专为AI而建", "★★★☆☆\n需导出", "活跃", "免费", "★★★☆☆\n学AI时代\nDS形态", "备选"],
    ["ui-ux-\ndesign-\nresources", "github.com/AyushWarrier/ui-ux-design-resources", "设计资源\n总集", "⭐经典", "UI/UX全栈资源\n20+灵感网站+Figma插件\n课程+作品集工具", "综合目录", "链接索引\nMarkdown", "★★☆☆☆\n资源链接\n需加工", "★★★★☆\n目录可做\n索引", "维护中", "免费", "★★☆☆☆\n工具+网站\n入口", "参考"],
]

row = 2
for d in dir2:
    is_sub = str(d[0]).startswith("═══")
    for c, v in enumerate(d[:13], 1):
        cl = ws2.cell(row=row, column=c, value=v)
        if c == 2 and not is_sub:
            cl.font = link_font
    style_row(ws2, row, 13, is_sub=is_sub)
    row += 1

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 30
for col in ['C','D','E','F','G','H','I','J','K','L','M']:
    ws2.column_dimensions[col].width = 20
ws2.freeze_panes = "C2"

# ═══════════════════════════════════════
# Sheet 3: 完整工作流方案
# ═══════════════════════════════════════
ws3 = wb.create_sheet("完整工作流方案")

ws3.merge_cells('A1:H1')
t1 = ws3.cell(row=1, column=1, value="从「捕获灵感」到「AI复刻」的完整工具链")
t1.font = Font(name="Arial", size=14, bold=True, color="2F5496")
t1.fill = sub_fill; t1.alignment = Alignment(horizontal="center")

ws3.merge_cells('A3:H3')
ws3.cell(row=3, column=1, value="你现有的架构 vs 增强后的架构").font = Font(name="Arial", size=11, bold=True)

ws3.merge_cells('A5:D5')
ws3.cell(row=5, column=1, value="现在（手动链路）").font = b_font; ws3['A5'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
now = [
    "1. 🔍 发现灵感网站 → 你手动截图",
    "2. 👁️ Gemini网页版 → 你手动上传截图、粘贴Prompt",
    "3. 📝 Obsidian → 手动复制Gemini输出、新建笔记",
    "4. 💻 Claude Code → 你手动指定MD文件路径让我读取",
    "5. ⚡ 生成代码 → 我按基因档案输出React代码",
]
for i, s in enumerate(now):
    ws3.merge_cells(f'A{6+i}:D{6+i}')
    ws3.cell(row=6+i, column=1, value=s).font = c_font

ws3.merge_cells('A12:D12')
ws3.cell(row=12, column=1, value="增强后（半自动链路）").font = b_font; ws3['A12'].fill = g_fill
new = [
    "1. 🔍 发现灵感 → design-bites/designlang 自动提取 → DESIGN.md",
    "2. 👁️ Gemini网页版 → 补充感性描述（视觉锚点层）",
    "3. 📝 合并输出 → 基因档案 = DESIGN.md + 你的感性笔记 → 存入Obsidian",
    "4. 💻 Claude Code → 我直接读Obsidian基因档案，零理解损耗",
    "5. ⚡ 生成代码 → 按精确的色值+DOM+动画参数生成",
]
for i, s in enumerate(new):
    ws3.merge_cells(f'A{13+i}:D{13+i}')
    ws3.cell(row=13+i, column=1, value=s).font = c_font

ws3.merge_cells('A19:H19')
ws3.cell(row=19, column=1, value="推荐工具组合").font = Font(name="Arial", size=12, bold=True, color="2F5496")

combo = [
    ["环节", "首选工具", "链接", "为什么", "备选"],
    ["提取网页\n设计系统", "design-bites\n(npx一键\n270+网站)", "npmjs.com/package/design-bites", "零依赖、按域名获取\n直接产出DESIGN.md\n跟awesome-design-md生态无缝", "distill-design"],
    ["DESIGN.md\n参考库", "awesome-design-md\n(VoltAgent\n91K⭐)", "github.com/VoltAgent/awesome-design-md", "58+品牌完整设计系统\n每份含9节标准结构\n这就是基因档案模板", "AsahinaMafuuyuu\n分支"],
    ["动画/交互\n提取", "designlang\n(--full\n--motion-runtime)", "github.com/Manavarya09/design-extract", "业界唯一双层动画捕获\n输出motion-tokens.json\n含编排+滚动检测", "Liftit\n(逐帧动画精度最高)"],
    ["UX模式\n学习", "UXonFly MCP\n+Raven MCP", "github.com/Phanikondru/uxonfly-mcp\ngithub.com/rhinocap/raven-mcp", "MCP服务器直接对接\nClaude Code实时查询\n设计规则AI原生", "ux-patterns-for-\ndevelopers"],
    ["灵感存储\n管理", "Obsidian\n(你已在用)", "obsidian.md", "纯Markdown=AI原生\n标签+双链+Dataview\n完全归你掌控", "ByteBox\n(如需代码片段管理)"],
    ["AI执行\n生成", "Claude Code\n+ DeepSeek API\n(你已在用)", "—", "读取Obsidian基因档案\n精准生成React代码\n不臆造不猜测", "你现在就在用"],
]
for i, d in enumerate(combo):
    r = 20 + i
    for c, v in enumerate(d, 1):
        cl = ws3.cell(row=r, column=c, value=v)
        cl.font = b_font if i == 0 else c_font
        cl.fill = sub_fill if i == 0 else PatternFill()
        cl.alignment = wrap; cl.border = thin
        if c == 3 and i > 0:
            cl.font = link_font

ws3.merge_cells('A28:H28')
ws3.cell(row=28, column=1, value="下一步行动（10分钟即可验证）").font = Font(name="Arial", size=12, bold=True, color="2F5496")

actions = [
    ["1", "安装 design-bites", "npm i -g design-bites", "2分钟", "270+网站一键DESIGN.md"],
    ["2", "测试提取", "design-bites add stripe.com\ncat stripe.com/DESIGN.md", "3分钟", "验证输出质量"],
    ["3", "克隆参考库", "git clone https://github.com/VoltAgent/awesome-design-md.git", "3分钟", "58+品牌基因档案模板"],
    ["4", "存入Obsidian", "把一份DESIGN.md拖入Obsidian\n加标签 #ui-dna", "2分钟", "验证兼容性"],
    ["5", "让我生成代码", "告诉我Obsidian里的文件路径\n我读出来生成React组件", "我来做", "跑通完整闭环"],
]
act_hdr = ["", "动作", "命令", "耗时", "产出"]
for c, t in enumerate(act_hdr, 1):
    cl = ws3.cell(row=29, column=c, value=t)
    cl.font = b_font; cl.fill = sub_fill; cl.alignment = center; cl.border = thin
for i, d in enumerate(actions):
    r = 30 + i
    for c, v in enumerate(d, 1):
        cl = ws3.cell(row=r, column=c, value=v)
        cl.font = b_font if c <= 2 else c_font
        cl.alignment = wrap; cl.border = thin

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 28
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 22
ws3.column_dimensions['F'].width = 16
ws3.column_dimensions['G'].width = 16
ws3.column_dimensions['H'].width = 16

# ═══════════════════════════════════════
# Sheet 4: 两大方向总览
# ═══════════════════════════════════════
ws4 = wb.create_sheet("两大方向总览")

ws4.merge_cells('A1:F1')
t4 = ws4.cell(row=1, column=1, value="方向一 vs 方向二 的定位关系")
t4.font = Font(name="Arial", size=13, bold=True, color="2F5496")
t4.fill = sub_fill; t4.alignment = Alignment(horizontal="center")

overview = [
    ["维度", "方向一：储存UI灵感", "方向二：学习UI结构", "两者关系"],
    ["核心问题", "「我看到的好东西放哪？」", "「怎么理解好在哪里？」", "方向二教怎么看\n方向一教怎么存"],
    ["关键工具", "design-bites + Obsidian\nClaude Code(生成端)", "awesome-design-md + UXonFly\ndistill-design", "方向二的输出\n存入方向一的库"],
    ["输出物", "基因档案(.md)→AI可执行代码", "设计系统知识→设计思维", "知识+数据=能力"],
    ["跟Obsidian", "就是Obsidian笔记本身\n纯Markdown原生存储", "可存Obsidian作为\n参考教材+模板库", "Obsidian是统一载体"],
    ["免费方案", "design-bites + Obsidian\n完全免费", "awesome-design-md MIT\nUXonFly MIT 全部免费", "零成本搭建"],
]
for i, d in enumerate(overview):
    r = 3 + i
    for c, v in enumerate(d, 1):
        cl = ws4.cell(row=r, column=c, value=v)
        cl.font = b_font if i == 0 else c_font
        cl.fill = sub_fill if i == 0 else PatternFill()
        cl.alignment = wrap; cl.border = thin

ws4.merge_cells('A10:F10')
ws4.cell(row=10, column=1, value="如果今天只能装一个工具，选什么？").font = Font(name="Arial", size=12, bold=True, color="C00000")

ws4.merge_cells('A11:F11')
c11 = ws4.cell(row=11, column=1, value="🥇 design-bites (npmjs.com/package/design-bites) — npx一键获取270+网站的完整设计系统，输出DESIGN.md直接拖入Obsidian。0配置0费用0学习曲线。")
c11.font = Font(name="Arial", size=11, bold=True); c11.alignment = left_wrap

ws4.merge_cells('A12:F12')
c12 = ws4.cell(row=12, column=1, value="🥈 awesome-design-md (github.com/VoltAgent/awesome-design-md) — 装不了CLI可直接GitHub下载ZIP，获得58+品牌的设计系统Markdown文件。")
c12.font = Font(name="Arial", size=11, bold=True); c12.alignment = left_wrap

ws4.merge_cells('A13:F13')
ws4.cell(row=13, column=1, value="两个可以同时用——design-bites负责「自动蒸馏你发现的网站」，awesome-design-md负责「参考大厂已有的分析」.").font = Font(name="Arial", size=10, color="666666")

ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 18

# ═══════════════════════════════════════
# Sheet 5: 所有链接汇总
# ═══════════════════════════════════════
ws5 = wb.create_sheet("所有链接汇总")

h5 = ["项目名称", "完整URL", "类型", "所属方向"]
for c, t in enumerate(h5, 1): ws5.cell(row=1, column=c, value=t)
style_hdr(ws5, 1, len(h5))

links = [
    # 方向一
    ["ByteBox", "https://github.com/nickxla/bytebox", "代码片段管理器", "方向一"],
    ["Seahorse", "https://github.com/SSBun/Seahorse", "macOS书签App", "方向一"],
    ["Faved", "https://github.com/denho/faved", "书签管理器", "方向一"],
    ["Bookmrk", "https://github.com/sarvan-2187/Bookmrk", "看板书签", "方向一"],
    ["Uiverse Galaxy", "https://uiverse.io | https://github.com/uiverse-io/galaxy", "社区UI库", "方向一"],
    ["Storybook v10", "https://github.com/storybookjs/storybook", "组件文档", "方向一"],
    ["Story UI", "https://github.com/southleft/story-ui", "AI故事生成器", "方向一"],
    ["ReUI", "https://github.com/keenthemes/ReUI", "shadcn模式库", "方向一"],
    ["Uikki", "https://www.npmjs.com/package/uikki", "CLI组件库", "方向一"],
    ["UI Syntax", "https://ui-syntax.com", "AI生成组件", "方向一"],
    ["Component Gallery", "https://component.gallery", "组件参考", "方向一"],
    ["StyleGallery", "https://github.com/changeroa/StyleGallery", "CSS模式集", "方向一"],

    # 方向二
    ["awesome-design-md (VoltAgent)", "https://github.com/VoltAgent/awesome-design-md", "DESIGN.md集合", "方向二"],
    ["awesome-design-md (AsahinaMafuuyuu)", "https://github.com/AsahinaMafuuyuu/awesome-design-md", "DESIGN.md社区分支", "方向二"],
    ["designmd.sh", "https://designmd.sh", "CLI注册表", "方向二"],
    ["TypeUI", "https://typeui.sh", "CLI创作+注册表", "方向二"],
    ["design-bites", "https://www.npmjs.com/package/design-bites", "CLI设计片段库", "方向二"],
    ["mddesign", "https://www.npmjs.com/package/mddesign", "全生命周期CLI", "方向二"],
    ["distill-design", "https://github.com/tototomato1457/distill-design", "网站蒸馏工具", "方向二"],
    ["design-resources (bradtraversy)", "https://github.com/bradtraversy/design-resources-for-developers", "设计资源大全", "方向二"],
    ["awesome-ui-libraries", "https://github.com/dalisoft/awesome-ui-libraries", "UI库大全", "方向二"],
    ["ux-patterns-for-developers", "https://github.com/thedaviddias/ux-patterns-for-developers", "UX模式文档", "方向二"],
    ["UXonFly MCP", "https://github.com/Phanikondru/uxonfly-mcp", "UX规则MCP", "方向二"],
    ["Raven MCP", "https://github.com/rhinocap/raven-mcp", "设计智能MCP", "方向二"],
    ["github-elements", "https://github.com/github/github-elements", "GitHub Web组件", "方向二"],
    ["facebook/astryx", "https://github.com/facebook/astryx", "Meta设计系统", "方向二"],
    ["Geeklego", "https://github.com/geekyants/geeklego", "AI原生设计系统", "方向二"],
    ["ui-ux-design-resources", "https://github.com/AyushWarrier/ui-ux-design-resources", "设计资源总集", "方向二"],

    # 提取工具（之前的搜索结果）
    ["designlang (design-extract)", "https://github.com/Manavarya09/design-extract", "设计系统提取CLI", "工具"],
    ["Liftit", "https://www.npmjs.com/package/@ahmedessyad/liftit", "动画提取CLI", "工具"],
    ["brandmd", "https://github.com/yuvrajangadsingh/brandmd", "品牌设计提取", "工具"],
    ["SkillUI", "https://www.npmjs.com/package/skillui", "设计系统提取(双模式)", "工具"],
    ["getdesign.md", "https://getdesign.md", "DESIGN.md浏览器", "工具"],
]

for i, d in enumerate(links):
    r = i + 2
    for c, v in enumerate(d, 1):
        cl = ws5.cell(row=r, column=c, value=v)
        if c == 2:
            cl.font = link_font
    style_row(ws5, r, 4)

ws5.column_dimensions['A'].width = 28
ws5.column_dimensions['B'].width = 52
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 12
ws5.freeze_panes = "A2"

# ===== 保存 =====
out = "/Users/yubo/Claude code test/UI灵感库工具深度分析.xlsx"
wb.save(out)
print(f"✅ {out}")
