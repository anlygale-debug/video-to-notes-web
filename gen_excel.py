import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "多模态API价格对比"

# --- Data ---
# (模型, 厂商, 输入¥, 输出¥, 输入$, 输出$, 缓存$, 上下文, 图片, 视频, 音频, 开源, 性价比评价, 备注, tier)
data = [
    # tier 1: 极致低价
    ("GLM-4.6V-Flash (9B)", "智谱", 0, 0, 0, 0, "—", "128K", "✅", "❌", "❌", "✅",
     "⭐⭐⭐ 完全免费！轻量图片理解首选，开源可商用", "9B参数，适合简单OCR/分类/打标", "green"),
    ("GPT-5 Nano", "OpenAI", 0.36, 2.9, 0.05, 0.40, "$0.02", "400K", "✅", "❌", "❌", "❌",
     "⭐⭐ 输入极便宜($0.05)，输出也低，适合批量预处理", "OpenAI生态，能力有限", "green"),
    ("Reka Edge (7B)", "Reka AI", 0.73, 0.73, 0.10, 0.10, "$0.085", "16K", "✅", "✅", "❌", "❌",
     "⭐⭐⭐ 输入输出同价$0.10，极轻量，适合高频简单任务", "7B参数，图片+视频", "green"),
    ("Qwen-VL-Plus", "阿里", 0.75, 2.25, 0.21, 0.63, "—", "131K", "✅", "❌", "❌", "❌",
     "⭐⭐⭐ 国内性价比王！中文OCR首选，¥0.75入/¥2.25出", "阿里百炼平台，新用户送7000万token", "green"),

    # tier 2: 性价比甜点
    ("Pixtral 12B", "Mistral", 1.1, 1.1, 0.15, 0.15, "—", "128K", "✅", "❌", "❌", "✅",
     "⭐⭐⭐ 最便宜有性能的视觉模型，双向$0.15统一价，开源", "Apache协议，可自部署", "blue"),
    ("GLM-4.6V (106B)", "智谱", 1, 3, 0.14, 0.42, "¥0.2", "128K", "✅", "❌", "❌", "✅",
     "⭐⭐⭐ ¥1入/¥3出+原生Function Call，国内低价优选", "开源106B MoE，32K内定价", "blue"),
    ("DeepSeek V4 Flash", "DeepSeek", 1, 2, 0.14, 0.28, "¥0.02", "1M", "✅", "❌", "❌", "❌",
     "⭐⭐⭐ 🏆性价比之王！谷段¥1/¥2，284B参数，识图超GPT-5.4", "⚠峰谷定价：白天¥2/¥4，周末全天谷段", "blue"),
    ("InternVL3 8B", "OpenGVLab", 1.5, 1.5, 0.20, 0.20, "—", "16K", "✅", "❌", "❌", "✅",
     "⭐⭐ 开源可自部署，适合内网场景，$0.20双向价", "8B参数，低成本自托管", "blue"),
    ("Grok 4.1 Fast", "xAI", 1.5, 3.6, 0.20, 0.50, "$0.05", "2M", "✅", "❌", "✅", "❌",
     "⭐⭐⭐ 极低价$0.20/$0.50，2M超长上下文，视觉+音频", "xAI出品，性价比炸裂", "blue"),
    ("Qwen-VL-Max", "阿里", 1.5, 4.5, 0.80, 3.20, "—", "131K", "✅", "❌", "❌", "❌",
     "⭐⭐ 复杂视觉理解强，¥1.5入/¥4.5出，百炼平台方便", "阿里云国内站定价", "blue"),
    ("Gemini 2.5 Flash", "Google", 2.2, 18.2, 0.30, 2.50, "—", "1M", "✅", "✅", "✅", "❌",
     "⭐⭐ Google原生多模态，生态好，输入极便宜", "输出偏贵", "blue"),
    ("GPT-5 Mini", "OpenAI", 1.8, 14.5, 0.25, 2.00, "—", "400K", "✅", "❌", "❌", "❌",
     "⭐⭐ OpenAI生态，品牌可靠，输入便宜", "输出中档水平", "blue"),
    ("MiniMax M3 (第三方)", "MiniMax", 2.2, 8.7, 0.30, 1.20, "$0.06", "≤512K", "✅", "✅", "❌", "❌",
     "⭐⭐⭐ 第三方价$0.30/$1.20比官方便宜一半！图片+视频", "Fireworks/OpenRouter渠道", "blue"),
    ("DeepSeek V4 Pro", "DeepSeek", 3, 6, 0.435, 0.87, "¥0.025", "1M", "✅", "❌", "❌", "❌",
     "⭐⭐⭐ 谷段¥3/¥6，1.6T参数旗舰性能，缓存仅¥0.025", "⚠峰谷定价：峰段¥6/¥12", "blue"),
    ("Step-3.7-Flash", "阶跃星辰", 1.35, 8.1, 0.20, 1.15, "¥0.27", "256K", "✅", "✅", "❌", "❌",
     "⭐⭐⭐ 198B MoE推理强，¥1.35入，图像+视频，订阅更省", "Step Plan ¥49/月起", "blue"),
    ("Doubao-Seed-2.1 Turbo", "字节", 3, 15, 0.41, 2.05, "—", "—", "✅", "✅", "❌", "❌",
     "⭐⭐ 豆包最新2.1 Turbo，高频调用场景，Pro版半价", "字节生态，适合国内调用", "blue"),
    ("Qwen3-VL-32B", "阿里", 3.8, 3.8, 0.52, 0.52, "—", "32K", "✅", "❌", "❌", "❌",
     "⭐⭐⭐ MoE高效架构，双向$0.52统一价，评测公认甜点", "阿里国际站定价", "blue"),
    ("Qwen3-Omni-30B", "阿里", 3.8, 3.8, 0.52, 0.52, "—", "32K", "✅", "✅", "✅", "❌",
     "⭐⭐⭐ 图+视频+音频全模态！$0.52统一价，全模态最值", "全模态理解，一网打尽", "blue"),

    # tier 3: 中档
    ("Hunyuan-Vision", "腾讯", 3, 9, 1.20, 1.20, "—", "32K", "✅", "❌", "❌", "❌",
     "⭐⭐ 国内渠道方便，腾讯云直接调用，适合企业用户", "免费额度100万token", "yellow"),
    ("GLM-5V-Turbo", "智谱", 5, 22, 1.20, 4.00, "$0.24", "200K", "✅", "✅", "❌", "❌",
     "⭐⭐⭐ 🏆多模态Agent首选！Design2Code 94.8分，比Opus便宜92%", "原生Function Call，设计稿转代码最强", "yellow"),
    ("InternVL3 38B/78B", "OpenGVLab", 6.6, 6.6, 0.90, 0.90, "$0.45", "16K", "✅", "❌", "❌", "✅",
     "⭐⭐ 38B/78B同价$0.90，开源可自部署，内网优选", "直接选78B更划算", "yellow"),
    ("Grok 4.3", "xAI", 9.1, 18.2, 1.25, 2.50, "$0.20", "1M", "✅", "❌", "❌", "❌",
     "⭐⭐⭐ 输出$2.50极便宜！1M上下文，xAI主力旗舰", "视觉+推理，输出价是同级最低", "yellow"),
    ("Kimi K2.6", "月之暗面", 6.5, 27, "—", "—", "—", "256K", "✅", "✅", "❌", "❌",
     "⭐⭐ 图文视频全支持，256K上下文，国内调用方便", "月之暗面出品", "yellow"),
    ("GPT-5", "OpenAI", 9.1, 73, 1.25, 10.00, "—", "272K", "✅", "❌", "❌", "❌",
     "⭐⭐ 综合能力强，生态最完善，但输出贵", "2025年8月发布，已有更新版", "yellow"),
    ("Gemini 3 Flash", "Google", 3.6, 21.8, 0.50, 3.00, "—", "1M", "✅", "✅", "✅", "❌",
     "⭐⭐ Google最新Flash，原生多模态，输入便宜输出中档", "Gemini 3系列最新", "yellow"),
    ("Claude Haiku 4.5", "Anthropic", 7.3, 36.4, 1.00, 5.00, "$0.10", "1M", "✅", "❌", "❌", "❌",
     "⭐⭐ 品质可靠速度快，1M上下文，缓存仅$0.10", "Anthropic出品，稳定可靠", "yellow"),
    ("Doubao-Seed-2.1 Pro", "字节", 6, 30, 0.82, 4.11, "¥1.2", "—", "✅", "✅", "❌", "❌",
     "⭐⭐ 豆包最新旗舰，Coding+Agent+VLM三合一", "字节生态最强", "yellow"),

    # tier 4: 高端旗舰
    ("Gemini 3.1 Pro", "Google", 14.6, 87.6, 2.00, 12.00, "—", "1M", "✅", "✅", "✅", "❌",
     "⭐ 原生多模态旗舰，1M上下文，不差钱首选Google", "Google最强多模态", "red"),
    ("GPT-5.4", "OpenAI", 18.2, 109, 2.50, 15.00, "$1.25", "1M+", "✅", "❌", "❌", "❌",
     "⭐ OpenAI最新旗舰，1M上下文，最强之一但最贵之一", "2026年3月发布", "red"),
    ("Pixtral Large (124B)", "Mistral", 14.6, 43.8, 2.00, 6.00, "—", "131K", "✅", "❌", "❌", "✅",
     "⭐ 开源124B视觉旗舰，输出$6相对便宜，可自部署", "Mistral最强视觉模型", "red"),
    ("Claude Sonnet 4.6", "Anthropic", 21.9, 109, 3.00, 15.00, "$0.30", "1M", "✅", "❌", "❌", "❌",
     "⭐ 编程+视觉综合最强，缓存$0.30极低，复杂推理首选", "Anthropic出品，品质标杆", "red"),
    ("Claude Opus 4.8", "Anthropic", 36.5, 182, 5.00, 25.00, "$0.50", "1M", "✅", "❌", "❌", "❌",
     "⭐ 最强但最贵，适合对精度要求极致的场景", "钱不是问题时选这个", "red"),
    ("MiniMax M3 (官方)", "MiniMax", 4.4, 17.5, 0.60, 2.40, "$0.12", "1M", "✅", "✅", "❌", "❌",
     "⭐ 428B MoE官方价，建议走第三方更便宜", "官方价比第三方贵一倍", "red"),
]

# --- Styles ---
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 极致低价
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')   # 性价比甜点
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 中档
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')     # 高端

green_font = Font(name='微软雅黑', size=10, color='006100', bold=True)
blue_font = Font(name='微软雅黑', size=10, color='1F4E79', bold=True)
yellow_font = Font(name='微软雅黑', size=10, color='9C6500', bold=True)
red_font = Font(name='微软雅黑', size=10, color='9C0006', bold=True)

tier_fills = {"green": green_fill, "blue": blue_fill, "yellow": yellow_fill, "red": red_fill}
tier_fonts = {"green": green_font, "blue": blue_font, "yellow": yellow_font, "red": red_font}
tier_labels = {"green": "💰 极致低价", "blue": "⭐ 性价比甜点", "yellow": "📊 中档", "red": "🏆 高端旗舰"}

# --- Headers ---
headers = [
    "排名", "模型名称", "厂商",
    "输入价格 ¥/1M", "输出价格 ¥/1M", "输入价格 $/1M", "输出价格 $/1M",
    "缓存输入", "上下文窗口",
    "图片", "视频", "音频", "开源",
    "💡 性价比评价", "备注"
]

# Row 1: Title
ws.merge_cells('A1:O1')
title_cell = ws['A1']
title_cell.value = "多模态 API 价格对比（2026年7月）— 20+ 模型全网搜罗，按性价比排序"
title_cell.font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ws.row_dimensions[1].height = 35

# Row 2: Legend
ws.merge_cells('A2:O2')
legend = ""
for key, label in tier_labels.items():
    legend += f"  {label}  "
ws['A2'].value = "色块图例：" + " | ".join([f"■ {v}" for v in tier_labels.values()])
ws['A2'].font = Font(name='微软雅黑', size=9, color='555555')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# Row 3: Headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[3].height = 30

# --- Data Rows ---
for i, row_data in enumerate(data):
    row_num = i + 4
    tier = row_data[-1]
    fill = tier_fills[tier]

    for col_idx, value in enumerate(row_data[:-1], 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

        # Color the tier label column (排名)
        if col_idx == 1:
            cell.font = tier_fonts[tier]
            cell.fill = fill
        else:
            cell.fill = fill

        # Left-align long text columns
        if col_idx in (2, 14, 15):
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.row_dimensions[row_num].height = 52

# --- Column Widths ---
col_widths = {
    'A': 5, 'B': 28, 'C': 12,
    'D': 14, 'E': 14, 'F': 16, 'G': 16,
    'H': 14, 'I': 12,
    'J': 7, 'K': 7, 'L': 7, 'M': 7,
    'N': 52, 'O': 36
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# --- Freeze panes ---
ws.freeze_panes = 'A4'

# --- Auto-filter ---
ws.auto_filter.ref = f"A3:O{3 + len(data)}"

# --- Sheet 2: Summary ---
ws2 = wb.create_sheet("推荐总结")

# Top picks
summary_data = [
    ("场景", "🏆 首选", "💰 价格", "💡 理由"),
    ("最低成本（简单OCR/打标）", "GLM-4.6V-Flash", "免费", "9B开源，API完全免费，轻量图片理解首选"),
    ("最低成本（有性能要求）", "Pixtral 12B", "$0.15/$0.15", "双向统一价$0.15，Apache开源可自部署"),
    ("中文OCR/文档理解", "Qwen-VL-Plus", "¥0.75/¥2.25", "阿里百炼平台，中文优化，新用户送7000万token"),
    ("综合性价比之王", "DeepSeek V4 Flash", "¥1/¥2 (谷段)", "284B参数，识图能力超GPT-5.4，周末全天谷段"),
    ("图片+视频+音频全模态", "Qwen3-Omni-30B", "$0.52/$0.52", "唯一全模态统一低价模型，30B MoE"),
    ("设计稿转代码", "GLM-5V-Turbo", "¥5/¥22", "Design2Code 94.8分，比Claude Opus便宜92%"),
    ("复杂视觉推理(中档)", "Grok 4.3", "$1.25/$2.50", "输出$2.50同级最低，1M上下文，xAI旗舰"),
    ("复杂推理不差钱", "Claude Sonnet 4.6", "$3/$15", "编程+视觉综合最强，品质标杆"),
    ("内网/隐私/自部署", "InternVL3 78B", "$0.90/$0.90", "开源可自部署，零API费用，78B性能足够"),
    ("视频理解", "MiniMax M3 (第三方)", "$0.30/$1.20", "OpenRouter渠道半价，原生支持图片+视频"),
    ("国内调用最方便", "Qwen-VL-Plus / GLM-4.6V", "¥0.75-1/3", "阿里/智谱国内平台，无需翻墙，中文最好"),
    ("高频批量调用", "DeepSeek V4 Flash", "缓存¥0.02/1M", "缓存命中几乎免费，284B参数，性价比无敌"),
]

for i, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=col_idx, value=value)
        cell.border = thin_border
        if i == 1:
            cell.font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left' if col_idx != 1 else 'center', vertical='center', wrap_text=True)
            if i % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    if i > 1:
        ws2.row_dimensions[i].height = 28

ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 26
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 52
ws2.freeze_panes = 'A2'

# --- Sheet 3: Key Notes ---
ws3 = wb.create_sheet("关键提示")

notes = [
    "📌 关键提示",
    "",
    "1. DeepSeek V4 峰谷定价",
    "   - 谷段（晚上、夜间、周末全天）：Flash ¥1/¥2，Pro ¥3/¥6",
    "   - 峰段（工作日 9:00-12:00, 14:00-18:00）：价格×2",
    "   - 建议：尽量错峰到晚上/周末调用，或使用缓存（命中仅¥0.02-0.025/1M）",
    "",
    "2. 第三方聚合平台更便宜",
    "   - MiniMax M3 官方 $0.60/$2.40 → Fireworks/OpenRouter $0.30/$1.20（便宜一半）",
    "   - 很多模型在 OpenRouter、Fireworks、Together 等平台有折扣",
    "   - 建议：先查第三方平台价格再决定从哪调用",
    "",
    "3. 缓存命中能省 80-90%",
    "   - DeepSeek V4 缓存输入 ¥0.02/1M（原价 ¥1/1M 的 2%）",
    "   - Claude 缓存输入 $0.30/1M（原价 $3/1M 的 10%）",
    "   - 高频重复 prompt（系统指令、多轮对话前缀）务必启用缓存",
    "",
    "4. 开源模型可自部署",
    "   - Pixtral 12B、InternVL3、GLM-4.6V 系列都开源",
    "   - 调用量大时自部署更省钱，只需 GPU 成本",
    "   - GLM-4.6V-Flash 甚至 API 就直接免费",
    "",
    "5. 货币换算参考",
    "   - 1 USD ≈ 7.3 RMB（2026年7月参考汇率）",
    "   - 国内模型（阿里、智谱、DeepSeek、阶跃、豆包）以人民币计价",
    "   - 国际模型（OpenAI、Anthropic、Google、xAI、Mistral）以美元计价",
    "",
    "6. 数据采集日期",
    "   - 所有价格采集于 2026年7月1日",
    "   - 大模型 API 价格变动频繁，使用前请以官方最新价格为准",
]

for i, note in enumerate(notes, 1):
    cell = ws3.cell(row=i, column=1, value=note)
    if i == 1:
        cell.font = Font(name='微软雅黑', bold=True, size=13, color='1F4E79')
    elif note.startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
        cell.font = Font(name='微软雅黑', bold=True, size=11)
    else:
        cell.font = Font(name='微软雅黑', size=10)

ws3.column_dimensions['A'].width = 80

# --- Save ---
output_path = "/Users/yubo/Claude code test/多模态API价格对比.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Sheet 1: {len(data)} models")
print(f"Sheet 2: {len(summary_data)-1} recommendations")
