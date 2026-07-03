#!/usr/bin/env python3
"""跨 4 站分析 + 用途相关性分级，生成 Excel 报告"""
import os, json, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
SITES = {
    "01-vercel": "SaaS 落地页",
    "02-linear": "Web App",
    "03-tailwindcss": "文档站",
    "04-stripe": "企业 SaaS",
}

# ========== 文件元信息：文件名 + 用途分类 + 对复刻的相关性 ==========
# 相关性: essential / useful / framework-specific / tooling / meta / irrelevant
FILE_META = {
    # ---- 设计 Token 层 ----
    "DESIGN.md": {
        "category": "设计 Token",
        "sub": "AI 原生设计文档",
        "relevance": "essential",
        "desc": "YAML 元数据 + 色值/字体/组件清单。给 AI 快速了解设计全貌。",
    },
    "design-language.md": {
        "category": "设计 Token",
        "sub": "19 节完整设计语言",
        "relevance": "essential",
        "desc": "最完整的设计文档，含色板 HSL、渐变原文、字重全量。AI 首选参考。",
    },
    "design-tokens.json": {
        "category": "设计 Token",
        "sub": "DTCG 标准 JSON",
        "relevance": "essential",
        "desc": "W3C DTCG 格式，primitive+semantic+composite 三层 token。机器可读。",
    },
    "variables.css": {
        "category": "设计 Token",
        "sub": "CSS 自定义属性",
        "relevance": "essential",
        "desc": "完整 CSS 变量：所有字号(64/56/30/…/8)、间距、色值、阴影。直接 import。",
    },
    "tailwind-v4.css": {
        "category": "设计 Token",
        "sub": "Tailwind v4 @theme",
        "relevance": "useful",
        "desc": "HSL 色阶(primary-50~950)+Tailwind 主题配置。用 Tailwind 时需要。",
    },
    "tokens.d.ts": {
        "category": "设计 Token",
        "sub": "TypeScript 类型",
        "relevance": "meta",
        "desc": "TypeScript 类型定义，非 TS 项目用不到。",
    },
    "tailwind.config.js": {
        "category": "设计 Token",
        "sub": "Tailwind v3 配置",
        "relevance": "meta",
        "desc": "v3 配置，不如 v4 的 CSS-first 好。有 v4 时不需要这个。",
    },
    "shadcn-theme.css": {
        "category": "设计 Token",
        "sub": "shadcn/ui 主题",
        "relevance": "framework-specific",
        "desc": "shadcn/ui globals.css。用 shadcn 时需要，否则不需要。",
    },
    "theme.js": {
        "category": "设计 Token",
        "sub": "React/Vue/Svelte 主题对象",
        "relevance": "useful",
        "desc": "框架无关的主题 JS 对象，可直接 import 用。",
    },
    "reset.css": {
        "category": "设计 Token",
        "sub": "品牌 CSS Reset",
        "relevance": "useful",
        "desc": "含 base font-size(16px)、font-family、body 色。从零建项目时有用。",
    },

    # ---- 动画/动效层 ----
    "motion.framer.js": {
        "category": "动画",
        "sub": "Framer Motion 预设",
        "relevance": "essential",
        "desc": "可直接 import 的 Framer Motion presets（transitions+variants+inView）。最常用。",
    },
    "motion.gsap.js": {
        "category": "动画",
        "sub": "GSAP 预设",
        "relevance": "framework-specific",
        "desc": "GSAP CustomEase+reveals+ScrollTrigger。用 GSAP 时需要。",
    },
    "motion.waapi.js": {
        "category": "动画",
        "sub": "Web Animations API",
        "relevance": "framework-specific",
        "desc": "零依赖 WAAPI 版本。不需要任何库就能用。备选方案。",
    },
    "motion.one.js": {
        "category": "动画",
        "sub": "Motion One 预设",
        "relevance": "framework-specific",
        "desc": "Motion One 库版本。小众，一般不需要。",
    },
    "motion.tailwind.js": {
        "category": "动画",
        "sub": "Tailwind 动画配置",
        "relevance": "framework-specific",
        "desc": "Tailwind animation 预设。用 Tailwind 动画时需要。",
    },
    "motion-tokens.json": {
        "category": "动画",
        "sub": "结构化动画参数",
        "relevance": "essential",
        "desc": "JSON 格式的 duration/easing/spring/feel 参数。机器可读的动画规范。",
    },
    "motion.css": {
        "category": "动画",
        "sub": "动画 CSS 变量+keyframes",
        "relevance": "useful",
        "desc": "CSS 动画变量和 @keyframes。纯 CSS 方案。",
    },
    "motion.html": {
        "category": "动画",
        "sub": "交互动画预览页",
        "relevance": "meta",
        "desc": "浏览器里看的动画预览页。给人演示用，不参与代码生成。",
    },

    # ---- 组件/结构层 ----
    "anatomy.tsx": {
        "category": "组件结构",
        "sub": "React 组件骨架",
        "relevance": "essential",
        "desc": "按钮/卡片等组件的 variant×size×state 矩阵 + React stub。复刻组件时必备。",
    },
    "intent.json": {
        "category": "页面结构",
        "sub": "页面意图+板块角色",
        "relevance": "essential",
        "desc": "页面类型(landing/pricing)+section 角色(hero/feature-grid/cta/footer)。重构页面结构必备。",
    },

    # ---- 品牌/风格层 ----
    "voice.json": {
        "category": "品牌语调",
        "sub": "语调+CTA+样本标题",
        "relevance": "essential",
        "desc": "tone/pronoun/heading style/CTA verbs/样本标题。AI 写文案时不跑偏。",
    },
    "visual-dna.json": {
        "category": "设计基因",
        "sub": "材质+图形+背景风格",
        "relevance": "essential",
        "desc": "flat/neumorphic/glass 材质、gradient-mesh 图形风格、饱和度、阴影画像。复刻「感觉」用。",
    },
    "gradients.css": {
        "category": "视觉素材",
        "sub": "CSS 渐变工具类",
        "relevance": "useful",
        "desc": "提取的渐变+工具类。站点有渐变时很重要，无渐变时为空。",
    },
    "gradients.json": {
        "category": "视觉素材",
        "sub": "渐变结构化数据",
        "relevance": "useful",
        "desc": "渐变的 JSON 版本。程序化使用。",
    },

    # ---- 截图/预览层 ----
    "preview.html": {
        "category": "截图预览",
        "sub": "可视化预览页",
        "relevance": "meta",
        "desc": "浏览器看的预览页。给人参考，不参与代码生成。",
    },
    "screenshots.json": {
        "category": "截图元数据",
        "sub": "截图索引",
        "relevance": "meta",
        "desc": "截图的索引文件。工具用，人不读。",
    },
    "responsive.json": {
        "category": "截图元数据",
        "sub": "响应式截图索引",
        "relevance": "meta",
        "desc": "响应式断点截图索引。工具用。",
    },

    # ---- 导出/平台层 ----
    "figma-variables.json": {
        "category": "导出",
        "sub": "Figma Variables",
        "relevance": "irrelevant",
        "desc": "Figma 插件导入用。和代码复刻无关。",
    },
    "wordpress-theme.json": {
        "category": "导出",
        "sub": "WordPress 主题",
        "relevance": "irrelevant",
        "desc": "WordPress 主题配置。不需要。",
    },
    "mcp.json": {
        "category": "工具元数据",
        "sub": "MCP Server 数据",
        "relevance": "meta",
        "desc": "MCP 协议数据。给 Claude Code 等 agent 自动加载用。不进 Obsidian。",
    },

    # ---- 诊断/元数据层 ----
    "AGENT.md": {
        "category": "AI 指令",
        "sub": "构建规则+输出期望",
        "relevance": "essential",
        "desc": "7 条构建规则(不造色值/间距对齐/匹配语调等)+输出期望。AI 代码生成的行为准则。",
    },
    "library.json": {
        "category": "诊断",
        "sub": "组件库检测",
        "relevance": "meta",
        "desc": "检测到 shadcn/ui / Radix 等。信息已在 DESIGN.md 里。",
    },
    "logo.json": {
        "category": "诊断",
        "sub": "Logo 元数据",
        "relevance": "irrelevant",
        "desc": "Logo URL+尺寸。和 UI 复刻无关。",
    },
    "icon-system.json": {
        "category": "诊断",
        "sub": "图标系统检测",
        "relevance": "meta",
        "desc": "检测到 SVG icon 的数量和类型。参考信息。",
    },
    "stack-intel.json": {
        "category": "诊断",
        "sub": "技术栈检测",
        "relevance": "irrelevant",
        "desc": "CMS/analytics/实验平台检测。和 UI 复刻无关。",
    },
    "seo.json": {
        "category": "诊断",
        "sub": "SEO 元数据",
        "relevance": "irrelevant",
        "desc": "meta tags/structured data。和 UI 复刻无关。",
    },
    "perf.json": {
        "category": "诊断",
        "sub": "性能数据",
        "relevance": "irrelevant",
        "desc": "资源大小/请求数。和 UI 复刻无关。",
    },
    "form-states.json": {
        "category": "诊断",
        "sub": "表单状态",
        "relevance": "meta",
        "desc": "表单验证状态。通常信息量少。",
    },
    "multipage.json": {
        "category": "诊断",
        "sub": "多页爬取结果",
        "relevance": "meta",
        "desc": "多页爬取的页面列表。单页提取时为空。",
    },
}

# ========== 分析逻辑 ==========

def find_file(site_dir, suffix):
    for f in os.listdir(site_dir):
        if f.endswith(suffix):
            return os.path.join(site_dir, f)
    return None


def file_summary(site_dir, suffix):
    """返回文件的简要统计"""
    path = find_file(str(site_dir), suffix)
    if not path:
        return {"存在": "❌", "大小": "—", "内容摘要": "文件不存在"}

    size = os.path.getsize(path)
    try:
        with open(path) as f:
            content = f.read()
    except:
        return {"存在": "✅", "大小": f"{size:,}B", "内容摘要": "(二进制)"}

    # 尝试读取内容摘要
    summary = ""
    if suffix.endswith(".json"):
        try:
            data = json.loads(content)
            top_keys = list(data.keys())[:5]
            summary = f"顶层键: {', '.join(top_keys)}"
        except:
            summary = f"JSON 解析失败"
    elif suffix.endswith(".js"):
        exports = re.findall(r"export (?:const|function|default) (\w+)", content)
        summary = f"导出: {', '.join(exports[:8])}" if exports else "无导出"
    elif suffix.endswith(".css"):
        vars_count = len(re.findall(r"--[\w-]+:", content))
        summary = f"CSS 变量: {vars_count} 个"
    elif suffix.endswith(".md"):
        sections = re.findall(r"^#{1,3}\s+(.+)$", content, re.MULTILINE)
        summary = f"章节: {', '.join(sections[:6])}"
    elif suffix.endswith(".tsx"):
        components = re.findall(r"export (?:function|interface) (\w+)", content)
        summary = f"组件: {', '.join(components[:5])}" if components else "空骨架"

    return {"存在": "✅", "大小": f"{size:,}B", "内容摘要": summary[:100]}


# ===== 生成 Excel =====
def generate_excel():
    wb = openpyxl.Workbook()

    # 样式
    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    essential_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    useful_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fw_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    meta_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    irrelevant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cell_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    relevance_fills = {
        "essential": essential_fill,
        "useful": useful_fill,
        "framework-specific": fw_fill,
        "meta": meta_fill,
        "irrelevant": irrelevant_fill,
    }
    relevance_labels = {
        "essential": "⚡ 必留",
        "useful": "✅ 有用",
        "framework-specific": "🔧 框架特定",
        "meta": "📋 元数据",
        "irrelevant": "❌ 无关",
    }

    # ===== Sheet 1: 综合分析 =====
    ws = wb.active
    ws.title = "综合分级"

    # 标题行
    headers = [
        "文件名", "用途分类", "子类别", "复刻相关性", "相关度说明",
        "vercel\n(SaaS 落地页)", "linear\n(Web App)",
        "tailwindcss\n(文档站)", "stripe\n(企业 SaaS)",
        "结论", "保留策略说明"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    row = 2
    # 按 relevance 排序输出
    order = {"essential": 0, "useful": 1, "framework-specific": 2, "meta": 3, "irrelevant": 4}
    sorted_files = sorted(FILE_META.items(), key=lambda x: order.get(x[1]["relevance"], 99))

    keep_count = 0
    skip_count = 0
    varies_count = 0

    for fname, meta in sorted_files:
        # 文件名
        ws.cell(row=row, column=1, value=fname).font = bold_font
        ws.cell(row=row, column=1).border = thin_border

        # 分类
        ws.cell(row=row, column=2, value=meta["category"]).font = cell_font
        ws.cell(row=row, column=2).border = thin_border

        # 子类
        ws.cell(row=row, column=3, value=meta["sub"]).font = cell_font
        ws.cell(row=row, column=3).border = thin_border

        # 相关性
        rel = meta["relevance"]
        rel_cell = ws.cell(row=row, column=4, value=relevance_labels.get(rel, rel))
        rel_cell.font = bold_font
        rel_cell.fill = relevance_fills.get(rel, PatternFill())
        rel_cell.alignment = center
        rel_cell.border = thin_border

        # 说明
        ws.cell(row=row, column=5, value=meta["desc"]).font = cell_font
        ws.cell(row=row, column=5).alignment = wrap
        ws.cell(row=row, column=5).border = thin_border

        # 4 站点数据
        for ci, (site_id, site_label) in enumerate(SITES.items()):
            site_dir = BASE / site_id
            summary = file_summary(site_dir, fname)
            text = f"{summary['存在']} {summary['大小']}\n{summary['内容摘要']}"
            cell = ws.cell(row=row, column=6 + ci, value=text)
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = thin_border
            if summary["存在"] == "❌":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # 结论
        if rel == "essential":
            conclusion = "✅ 固定保留"
            keep_count += 1
        elif rel == "irrelevant":
            conclusion = "❌ 固定跳过"
            skip_count += 1
        elif rel in ("meta", "framework-specific"):
            conclusion = "📋 固定跳过\n(参考/诊断用)"
            skip_count += 1
        else:
            conclusion = "🔶 按内容判断"
            varies_count += 1

        conc_cell = ws.cell(row=row, column=10, value=conclusion)
        conc_cell.font = bold_font
        conc_cell.alignment = center
        conc_cell.border = thin_border
        if "保留" in conclusion:
            conc_cell.fill = essential_fill
        elif "跳过" in conclusion:
            conc_cell.fill = irrelevant_fill

        # 策略说明
        if rel == "essential":
            strategy = "始终复制到 Obsidian"
        elif rel == "irrelevant":
            strategy = "始终跳过"
        elif rel in ("meta", "framework-specific"):
            strategy = "始终跳过（不进 Obsidian）"
        else:
            strategy = "内容丰富时复制，空时跳过"
        ws.cell(row=row, column=11, value=strategy).font = cell_font
        ws.cell(row=row, column=11).border = thin_border

        row += 1

    # 列宽
    widths = [22, 14, 20, 14, 40, 28, 28, 28, 28, 18, 24]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "F2"

    # 汇总行
    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    summary_cell = ws.cell(row=row, column=1, value=f"汇总：{keep_count} 个固定保留 + {varies_count} 个按内容判断 + {skip_count} 个固定跳过 = {len(FILE_META)} 个文件")
    summary_cell.font = Font(name="Arial", size=12, bold=True, color="2F5496")

    # ===== Sheet 2: 保留策略速查 =====
    ws2 = wb.create_sheet("保留策略速查")

    ws2.merge_cells("A1:C1")
    ws2.cell(row=1, column=1, value="Obsidian 灵感库 — 文件保留策略").font = Font(name="Arial", size=14, bold=True, color="2F5496")

    # 固定保留
    ws2.cell(row=3, column=1, value="✅ 固定保留（8 个）").font = Font(name="Arial", size=11, bold=True, color="006100")
    essential_files = [f for f, m in FILE_META.items() if m["relevance"] == "essential"]
    for i, f in enumerate(essential_files):
        ws2.cell(row=4 + i, column=1, value=f).font = bold_font
        ws2.cell(row=4 + i, column=2, value=FILE_META[f]["sub"]).font = cell_font
        ws2.cell(row=4 + i, column=3, value=FILE_META[f]["desc"]).font = cell_font

    # 按内容判断
    row2 = 4 + len(essential_files) + 1
    ws2.cell(row=row2, column=1, value="🔶 按内容判断（5 个）").font = Font(name="Arial", size=11, bold=True, color="9C6500")
    useful_files = [f for f, m in FILE_META.items() if m["relevance"] == "useful"]
    for i, f in enumerate(useful_files):
        ws2.cell(row=row2 + 1 + i, column=1, value=f).font = bold_font
        ws2.cell(row=row2 + 1 + i, column=2, value=FILE_META[f]["sub"]).font = cell_font
        ws2.cell(row=row2 + 1 + i, column=3, value=FILE_META[f]["desc"]).font = cell_font

    # 固定跳过
    n_useful = len(useful_files)
    row3 = row2 + 1 + n_useful + 1
    ws2.cell(row=row3, column=1, value="❌ 固定跳过（27 个）").font = Font(name="Arial", size=11, bold=True, color="9C0006")
    skip_files = [f for f, m in FILE_META.items() if m["relevance"] in ("framework-specific", "meta", "irrelevant")]
    for i, f in enumerate(skip_files):
        ws2.cell(row=row3 + 1 + i, column=1, value=f).font = bold_font
        ws2.cell(row=row3 + 1 + i, column=2, value=FILE_META[f]["sub"]).font = cell_font
        ws2.cell(row=row3 + 1 + i, column=3, value=FILE_META[f]["desc"]).font = cell_font

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 60

    # ===== Sheet 3: 站点间差异对比 =====
    ws3 = wb.create_sheet("站点间差异")
    headers3 = ["文件", "分类", "vercel", "linear", "tailwindcss", "stripe", "差异说明"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

    row = 2
    for fname, meta in sorted_files:
        if meta["relevance"] == "irrelevant":
            continue

        ws3.cell(row=row, column=1, value=fname).font = bold_font
        ws3.cell(row=row, column=1).border = thin_border
        ws3.cell(row=row, column=2, value=meta["sub"]).font = cell_font
        ws3.cell(row=row, column=2).border = thin_border

        sizes = []
        for ci, (site_id, site_label) in enumerate(SITES.items()):
            site_dir = BASE / site_id
            path = find_file(str(site_dir), fname)
            size = os.path.getsize(path) if path else 0
            sizes.append(size)
            text = f"{size:,}B" if size > 0 else "—"
            cell = ws3.cell(row=row, column=3 + ci, value=text)
            cell.font = cell_font
            cell.alignment = center
            cell.border = thin_border

        # 差异说明
        if max(sizes) > 0 and min(sizes) > 0:
            ratio = max(sizes) / max(min(sizes), 1)
            if ratio > 5:
                diff = f"⚠ 内容量差异大（{ratio:.0f}x）"
            elif ratio > 2:
                diff = f"🔶 有差异（{ratio:.0f}x）"
            else:
                diff = "✅ 内容量相近"
        elif max(sizes) > 0:
            diff = "⚠ 部分站点无此文件"
        else:
            diff = "—"

        ws3.cell(row=row, column=7, value=diff).font = cell_font
        ws3.cell(row=row, column=7).border = thin_border

        if "⚠" in diff:
            ws3.cell(row=row, column=7).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        row += 1

    widths3 = [22, 20, 14, 14, 14, 14, 30]
    for ci, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "C2"

    # 保存
    outpath = BASE / "UI灵感库_文件保留策略分析.xlsx"
    wb.save(outpath)
    print(f"Excel 已保存: {outpath}")
    print(f"\n=== 最终结论 ===")
    print(f"固定保留: {len(essential_files)} 个")
    for f in essential_files:
        print(f"  ✅ {f} — {FILE_META[f]['sub']}")
    print(f"\n按内容判断: {len(useful_files)} 个")
    for f in useful_files:
        print(f"  🔶 {f} — {FILE_META[f]['sub']}")
    print(f"\n固定跳过: {len(skip_files)} 个")
    print(f"  包括框架特定 {sum(1 for f,m in FILE_META.items() if m['relevance']=='framework-specific')} + 元数据 {sum(1 for f,m in FILE_META.items() if m['relevance']=='meta')} + 无关 {sum(1 for f,m in FILE_META.items() if m['relevance']=='irrelevant')}")


if __name__ == "__main__":
    generate_excel()
