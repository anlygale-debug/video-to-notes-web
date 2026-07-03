#!/usr/bin/env python3
"""GitHub 最热门 Skill Top 10 — Excel 表格"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATE = datetime.now().strftime("%Y-%m-%d")
FILENAME = f"热门Skills_Top10_{DATE}.xlsx"

SKILLS = [
    {
        "rank": 1,
        "name": "agent-reach",
        "source": "panniantong/agent-reach",
        "installs": 4700,
        "category": "全网搜索",
        "desc": "给你的 AI Agent 装上眼睛，看遍整个互联网。17 个平台一站式搜索（Twitter/X、Reddit、YouTube、GitHub、B站、小红书、微博、抖音、V2EX等），支持 CLI、MCP、curl、Python 四种调用方式，8 个平台零配置即开即用。无需 API Key，不产生费用。自带路由表，按场景自动选最优方案。",
        "stars": 19798,
        "pushed": "2026-05-18",
        "risk": "低风险",
    },
    {
        "rank": 2,
        "name": "browser-automation",
        "source": "sophieguanongit/openclaw-browser-automation",
        "installs": 2700,
        "category": "浏览器自动化",
        "desc": "基于 Playwright 的浏览器自动化控制。支持 CDP 模式连接已有 Chrome 或启动新 Chromium，复用页面和 Cookie，持久化登录状态。可导航、点击、填表单、截图、获取内容、等待元素、执行 JavaScript，覆盖浏览器所有交互操作。",
        "stars": 0,
        "pushed": "2026-02-18",
        "risk": "中风险",
    },
    {
        "rank": 3,
        "name": "game-development",
        "source": "sickn33/antigravity-awesome-skills",
        "installs": 2400,
        "category": "游戏开发",
        "desc": "安装量 1400+ Agentic Skills 的 GitHub 开源库，涵盖 Claude Code、Cursor、Codex CLI、Gemini CLI、Antigravity 等主流 AI 编程平台。包含安装器 CLI、技能包、工作流模板、官方及社区技能合集，是最全面的 Agent 技能市场之一。game-development 是其游戏开发子技能。",
        "stars": 37893,
        "pushed": "2026-05-18",
        "risk": "中风险",
    },
    {
        "rank": 4,
        "name": "douyin-video-summary",
        "source": "liu-wei-ai/douyin-video-summary",
        "installs": 1200,
        "category": "视频分析",
        "desc": "抖音视频智能总结工具。通过浏览器拦截音频请求提取音频 → ffmpeg 转 WAV → whisper.cpp 本地 Whisper 模型转录 → AI 生成结构化摘要（核心观点 + 要点列表 + 一句话总结）。可选同步到飞书文档。Mac 专用，Metal GPU 加速推理。",
        "stars": 0,
        "pushed": "2026-02-10",
        "risk": "低风险",
    },
    {
        "rank": 5,
        "name": "game-design-theory",
        "source": "pluginagentmarketplace/custom-plugin-game-developer",
        "installs": 1200,
        "category": "游戏设计",
        "desc": "游戏开发者路线图插件，内置引擎专用设计模式与优化策略。涵盖关卡设计、游戏机制建模、玩法循环设计、难度曲线调优等游戏设计核心理论，适用于 Unity/Unreal/Godot 等多引擎场景。",
        "stars": 22,
        "pushed": "2026-01-05",
        "risk": "低风险",
    },
    {
        "rank": 6,
        "name": "create-prd",
        "source": "phuryn/pm-skills",
        "installs": 918,
        "category": "产品管理",
        "desc": "PM Skills 市场的核心技能——用 8 章节标准模板撰写产品需求文档（PRD）。覆盖问题定义、目标指标、用户分群、价值主张、解决方案、发布计划等全流程。属于 phuryn/pm-skills（100+ Agentic Skills 的 PM 技能市场，11,355 GitHub Stars）。",
        "stars": 11355,
        "pushed": "2026-04-22",
        "risk": "低风险",
    },
    {
        "rank": 7,
        "name": "web-search",
        "source": "jwynia/agent-skills",
        "installs": 667,
        "category": "网页搜索",
        "desc": "通用网页搜索引擎 Skill，支持多引擎聚合搜索，将互联网信息转化为结构化数据。适用于资料收集、竞品调研、技术文档查找等场景，搜索结果可与其他 Skill 联动做进一步处理。",
        "stars": 79,
        "pushed": "2026-02-24",
        "risk": "低风险",
    },
    {
        "rank": 8,
        "name": "chrome-automation",
        "source": "zc277584121/marketing-skills",
        "installs": 630,
        "category": "浏览器自动化",
        "desc": "营销自动化技能包中的 Chrome 自动化模块。AI 接管浏览器执行重复操作：打开网页、点击按钮、填写表单、截取截图、抓取数据，适用于竞品监控、数据采集、自动签到等营销场景。",
        "stars": 0,
        "pushed": "2026-05-09",
        "risk": "中风险",
    },
    {
        "rank": 9,
        "name": "game-designer",
        "source": "opusgamelabs/game-creator",
        "installs": 618,
        "category": "游戏设计",
        "desc": "Opus Game Labs 出品的游戏创作工具，支持 2D（Phaser 引擎）和 3D（Three.js 引擎）游戏开发。从游戏概念到可运行原型的一条龙工作流，内置关卡编辑器、精灵管理、物理引擎集成、音效系统等模块。",
        "stars": 160,
        "pushed": "2026-05-12",
        "risk": "低风险",
    },
    {
        "rank": 10,
        "name": "ffmpeg-analyse-video",
        "source": "fabriqaai/ffmpeg-analyse-video-skill",
        "installs": 488,
        "category": "视频分析",
        "desc": "结合 ffmpeg 帧提取与 AI 视觉模型分析视频内容。自动提取关键帧 → AI 视觉识别画面 → 生成带时间戳的结构化摘要。支持任意格式视频文件，无需网络字幕，纯本地视觉分析路径，适合长视频内容理解和归档。",
        "stars": 10,
        "pushed": "2026-02-15",
        "risk": "低风险",
    },
]


def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills Top 10"

    # 列定义
    cols = [
        ("排名", 6),
        ("Skill 名称", 22),
        ("来源 (owner/repo)", 40),
        ("安装量", 10),
        ("GitHub Stars", 13),
        ("分类", 14),
        ("风险评级", 10),
        ("最近更新", 13),
        ("详细描述", 70),
    ]

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1,
                         value=f"GitHub Agent Skills 热门 Top 10 — {DATE}")
    title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="1A3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 副标题
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    sub_cell = ws.cell(row=2, column=1,
                       value="数据来源: npx skills find + GitHub API | 按安装量降序排列")
    sub_cell.font = Font(name="微软雅黑", size=9, color="888888")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # 表头
    hf = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tb = Border(left=Side("thin"), right=Side("thin"),
                top=Side("thin"), bottom=Side("thin"))

    for c, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=name)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = tb
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[3].height = 24

    # 数据行
    df = Font(name="微软雅黑", size=10)
    da = Alignment(vertical="center", wrap_text=True)
    dc = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")

    # 风险评级颜色
    risk_colors = {
        "低风险": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "中风险": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "高风险": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }

    for i, s in enumerate(SKILLS):
        row = 4 + i
        vals = [
            s["rank"], s["name"], s["source"], s["installs"],
            s["stars"], s["category"], s["risk"], s["pushed"], s["desc"]
        ]

        # 排名特殊颜色
        rank_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid") if s["rank"] <= 3 else None

        for j, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.font = df
            cell.border = tb
            if j in (1, 4, 5, 6, 7, 8):  # 居中列
                cell.alignment = dc
            else:
                cell.alignment = da

            # 排名高亮
            if j == 1 and rank_fill:
                cell.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
                cell.fill = rank_fill

            # 风险评级颜色
            if j == 7:
                rf = risk_colors.get(val)
                if rf:
                    cell.fill = rf
                    cell.font = Font(name="微软雅黑", size=10, bold=True)

            # 斑马纹
            if i % 2 == 1 and j != 7 and not (j == 1 and rank_fill):
                cell.fill = alt_fill

        ws.row_dimensions[row].height = 72

    # 冻结表头
    ws.freeze_panes = "A4"

    # 筛选器
    ws.auto_filter.ref = f"A3:I{4 + len(SKILLS) - 1}"

    # 底部汇总
    summary_row = 4 + len(SKILLS) + 1
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=len(cols))
    total_installs = sum(s["installs"] for s in SKILLS)
    total_stars = sum(s["stars"] for s in SKILLS)
    sc = ws.cell(row=summary_row, column=1,
                 value=f"汇总: Top 10 合计安装量 {total_installs:,} | GitHub Stars 合计 {total_stars:,} | 数据采集时间 {DATE}")
    sc.font = Font(name="微软雅黑", size=10, italic=True, color="666666")
    sc.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(FILENAME)
    return FILENAME


if __name__ == "__main__":
    fname = build_excel()
    print(f"已生成: {fname}")
    print(f"Top 10 技能涵盖: 搜索 | 浏览器自动化 | 游戏开发 | 视频分析 | 产品管理")
